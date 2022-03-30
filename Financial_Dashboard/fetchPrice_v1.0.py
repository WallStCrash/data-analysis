'''
Title:     fetchPrice_v1.0.py
Author:    wsc  
Date:      28.03.2022
    
Purpose:   This is a Python script to automate fetching prices for Equities, Bonds, FX, Commodities.
    
Notes:     Keep any eye on the URL of the item you are looking at as they can potentially change...
           Write a PowerShell script to run every weekday and then each day you will have a
           list of all the values as needed, accurate as of the close the previous day. Can then
           use Windows Task Scheduler to the below PowerShell script (or similar batch script) to 
           run every weekday!

PowerShell Script automate_price_fetch.ps1 is along the lines of:
           cd working_dir_containing_this_script
           python fetchPrice_v1.0.py
'''

# Import libraries and dependencies
import pandas as pd
import yfinance as yf
import pandas as pd
import openpyxl

from requests_html import HTMLSession
from pathlib import Path
from datetime import date

# Step 1: Get the information on indices
# Initialise an empty dictionary called "indicies"
# Create function to get the latest close value for that index using the <yfinance> package
# Use the index's ticker symbol to obtain the value and add to dictionary
indices = {}
def get_index(ticker):
        value_begin = yf.download(ticker, period="1d")
        value_final = value_begin["Close"][0]
        indices[ticker] = value_final
tickers = ["^FTSE", "^FTMC", "^GSPC","^STOXX", "^N225"]
for ticker in tickers:
        get_index(ticker)


# Step 2: Get the information on FX pairs
# Initialise an empty dictionary called "currencies"
# Create a function to get latest "close" value, not really close as FX trades 24/5.
# This function also utilises the <yfinance> package
# Use fx_pairs as the to run over and add to dictionary
currencies = {}
def get_fx(fx_pair):
        value_begin = yf.download(fx_pair, period="1d")
        value_final = value_begin["Close"][0]
        currencies[fx_pair] = value_final
fx_pairs = ["EURUSD=X" , "USDJPY=X" ,"GBPUSD=X" ,"AUDUSD=X" , "USDCAD=X" ,"USDCHF=X"]
for fx_pair in fx_pairs:
        get_fx(fx_pair)



# Step 3: Get BOND values
# Initialise an empty dictionary called "bonds"
# Create a function using the requests_html module as not possible to get individual security prices for bonds on <yfinance>
# We will obtain our values from a reputable Reuters Markets webpage
bonds = {}
def get_price_bond(url):
    s = s = HTMLSession()
    r = s.get(url)
    r.html.render(sleep = 1)
    value0 = (r.html.xpath("/html/body/div[2]/div[3]/section[2]/div[2]/div[2]/div[3]/div/div[2]/table/tbody/tr[2]/td[3]/span", first=True).text)
    bonds["USD-10y"] = value0
    value1 = (r.html.xpath("/html/body/div[2]/div[3]/section[2]/div[2]/div[2]/div[3]/div/div[2]/table/tbody/tr[3]/td[3]/span", first=True).text)
    bonds["UK-10y"] = value1
    value2 = (r.html.xpath("/html/body/div[2]/div[3]/section[2]/div[2]/div[2]/div[3]/div/div[2]/table/tbody/tr[10]/td[3]/span", first=True).text)
    bonds["DE-10y"] = value2
get_price_bond("https://www.wsj.com/market-data/quotes/bond/BX/TMBMKGB-10Y")



# Step 4: Get COMMODITIES  values using the same html methodlogy
# Create a function using the requests_html module as not possible to get individual commodity prices for bonds on <yfinance>
# We will obtain our values from a reputable Reuters Markets webpage
commodities = {}
def get_commodities_price():
    s = s = HTMLSession()
    r = s.get("https://www.wsj.com/market-data/quotes/futures/UK/IFEU/BRN00?mod=md_cmd_overview_quote")
    r.html.render(sleep = 1)
    value0 = (r.html.xpath("/html/body/div[2]/section[1]/div[2]/div/div[1]/ul[1]/li[2]/span/span[1]/span", first=True).text)
    commodities["BrentCrude"] = value0
    r = s.get("https://www.wsj.com/market-data/quotes/futures/GC00?mod=md_cmd_overview_quote")
    r.html.render(sleep = 1)
    value1 = (r.html.xpath("/html/body/div[2]/section[1]/div[2]/div/div[1]/ul[1]/li[2]/span/span[1]/span", first=True).text)
    commodities["Gold"] = value1
get_commodities_price()



# Test during development: let's see if indeed dictionaries have been created and populated correctly
# If we find spreadsheet not running properly in future, return here and check the output versus manually
# collated values to see where it is going wrong in the fetching prices stage or, if it is something with
# the excel stage...
print("Index info\n", indices)
print("FX info\n", currencies)
print("Bond info\n", bonds)
print("Commodity info\n", commodities)

# Step 5: Populate the Microsoft Excel workbook we use to get a daily update
root = Path("C:\\pathfile_to_where_excel_workbook_is_saved")
book = openpyxl.load_workbook(root/"Strategy_workbook.xlsx",data_only=True)
sheet = book.active
today = date.today()

# Timestamp our "updated" cell in excel so we can see when programme last run (automated)
sheet.cell(row=2, column=3, value=today.strftime("%d/%m/%Y"))

# Strip out forward slashes and timestamp our archive copy for future reference
date_as_string = today.strftime("%d%m%Y")

# This is the title we will save our archive copy as for future reference
title = f"C:pathfile_to_original_file\\StrategyWB_{date_as_string}.xlsx"

# Add values to the indices section
sheet.cell(row=5, column=3, value=indices["^FTSE"])
sheet.cell(row=6, column=3, value=indices["^FTMC"])
sheet.cell(row=7, column=3, value=indices["^GSPC"])
sheet.cell(row=8, column=3, value=indices["^STOXX"])
sheet.cell(row=9, column=3, value=indices["^N225"])

# Add values to FX section
sheet.cell(row=13, column=3, value=currencies["EURUSD=X"])
sheet.cell(row=14, column=3, value=currencies["USDJPY=X"])
sheet.cell(row=15, column=3, value=currencies["GBPUSD=X"])
sheet.cell(row=16, column=3, value=currencies["AUDUSD=X"])
sheet.cell(row=17, column=3, value=currencies["USDCAD=X"])
sheet.cell(row=18, column=3, value=currencies["USDCHF=X"])

#Add values to Bonds section
sheet.cell(row=21, column=3, value=bonds["USD-10y"])
sheet.cell(row=22, column=3, value=bonds["UK-10y"])
sheet.cell(row=23, column=3, value=bonds["DE-10y"])

# Add values to Commodities
sheet.cell(row=5, column=6, value=commodities["BrentCrude"])
sheet.cell(row=6, column=6, value=commodities["Gold"])

# Save workbook twice, overwrite the latest instance and once for the archive
book.save("C:\\original_pathfile.xlsx") # overwrite original
book.save(title) # save an archive copy

# Exit the session
exit()

